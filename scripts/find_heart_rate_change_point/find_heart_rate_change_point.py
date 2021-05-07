import xlrd, os, time, math
from openpyxl import Workbook, load_workbook
from string import digits

def main():
    run_test_file()

def run_test_file(input_file = "男生十周心率.xlsx"):
    input_wb = load_workbook(input_file)
    input_ws = input_wb.active
    name_row = 1
    while input_ws.cell(row=name_row, column=2).value is not None:
        name = input_ws.cell(row=name_row, column=2).value
        print(name)
        name_row += 1

if __name__ == "__main__":
    main()
