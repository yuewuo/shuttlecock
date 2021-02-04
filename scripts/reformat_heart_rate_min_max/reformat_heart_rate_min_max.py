import xlrd, os, time, math, sys
from openpyxl import Workbook, load_workbook
from string import digits
sys.path.append(os.path.realpath('../wechat_parse'))
from wechat_parse import parse_txt_chat

classes = ["class3", "class4", "class5"]

def main():
    for class_name in classes:
        reformat_heart_rate_min_max(class_name + ".xlsx", "../analyze_heart_rate/data_" + class_name + ".xlsx",
            parse_txt_chat("../wechat_parse/" + class_name + ".txt"), "../match_psycho_data/cleandata_all.xlsx",
            "reformat_min_max_" + class_name + ".xlsx")
        # exit(0)

def reformat_heart_rate_min_max(input_file, heart_rate_file, parsed_data, cleandata_file, output_file):
    input_wb = load_workbook(input_file)
    input_ws = input_wb.active
    assert input_ws.cell(row=1, column=2).value == "姓名"
    assert input_ws.cell(row=1, column=3).value == "年龄"
    assert input_ws.cell(row=1, column=4).value == "年级"
    assert input_ws.cell(row=1, column=14).value == "日期"
    assert input_ws.cell(row=1, column=15).value == "平均心率"
    assert input_ws.cell(row=1, column=16).value == "最大值"
    assert input_ws.cell(row=1, column=17).value == "最小值"
    assert input_ws.cell(row=1, column=18).value == "标准差"
    assert input_ws.cell(row=1, column=19).value == "三分钟总数"
    assert input_ws.cell(row=1, column=20).value == "一次连续最多"
    assert input_ws.cell(row=1, column=21).value == "感觉"
    # build name mapping
    name_row = 2
    name_to_row_mapping = { }
    id_to_row_mapping = { }
    while input_ws.cell(row=name_row, column=2).value is not None:
        name = input_ws.cell(row=name_row, column=2).value
        name_to_row_mapping[name] = name_row
        student_id = str(input_ws.cell(row=name_row, column=1).value)
        id_to_row_mapping[student_id] = name_row
        name_row += 1
    name_none_row = name_row
    # find first none column
    first_none_column = 3
    while input_ws.cell(row=1, column=first_none_column).value is not None:
        first_none_column += 1
    heart_rate_wb = load_workbook(heart_rate_file)
    sheet_length = len(heart_rate_wb.sheetnames)
    get_name_new_row = lambda row: (row - 2) * sheet_length + 2
    while name_row >= 2:
        name_new_row = get_name_new_row(name_row)
        for i in range(sheet_length):
            for j in range(1, first_none_column):
                if i == 0:
                    input_ws.cell(row=name_new_row+i, column=j).value = input_ws.cell(row=name_row, column=j).value
                else:
                    input_ws.cell(row=name_new_row+i, column=j).value = None
        name_row -= 1
    date_idx_mapping = {}
    for (heart_rate_idx, sheet_name) in enumerate(heart_rate_wb.sheetnames):
        date_idx_mapping[sheet_name] = heart_rate_idx
    for name_row in range(2, name_none_row):
        name_new_row = get_name_new_row(name_row)
        for sheet_name in heart_rate_wb.sheetnames:
            heart_rate_idx = date_idx_mapping[sheet_name]
            input_ws.cell(row=name_new_row + heart_rate_idx, column=14).value = sheet_name
    # mapping parsed_data
    remap_parsed_data = {}
    for (tag, mapping) in parsed_data:
        date = tag.split("/")
        assert len(date) == 3 and date[0] == "2020", "tag format must be like 2020/10/22"
        remap_date = date[1] + "." + date[2]
        remap_parsed_data[remap_date] = mapping
        if remap_date not in date_idx_mapping:
            print("[warning] wechat has date %s but not appear in original file %s" % (remap_date, input_file))
    # load data from heart rate file
    for sheet_name in heart_rate_wb.sheetnames:
        heart_rate_idx = date_idx_mapping[sheet_name]
        heart_rate_ws = heart_rate_wb[sheet_name]
        assert heart_rate_ws.cell(row=1, column=2).value == "秒数"
        a_name_row = 2
        while heart_rate_ws.cell(row=a_name_row, column=1).value is not None:
            name = heart_rate_ws.cell(row=a_name_row, column=1).value.translate(str.maketrans('', '', digits))  # delete numbers in name
            if name in name_to_row_mapping:
                name_row = name_to_row_mapping[name]
                name_new_row = get_name_new_row(name_row)
                person_data = []
                for i in range(180):
                    value = heart_rate_ws.cell(row=a_name_row, column=3+i).value
                    if value is not None:
                        person_data.append(int(value))
                # compute min max dev
                person_min = min(person_data)
                person_max = max(person_data)
                average = float(sum(person_data)) / len(person_data)
                stddev = math.sqrt(sum([(e - average) ** 2 for e in person_data]) / len(person_data))
                # print(name, person_min, person_max, stddev)
                input_ws.cell(row=name_new_row + heart_rate_idx, column=15).value = average
                input_ws.cell(row=name_new_row + heart_rate_idx, column=16).value = person_max
                input_ws.cell(row=name_new_row + heart_rate_idx, column=17).value = person_min
                input_ws.cell(row=name_new_row + heart_rate_idx, column=18).value = stddev
                if sheet_name in remap_parsed_data:
                    wechat_mapping = remap_parsed_data[sheet_name]
                    if name in wechat_mapping:
                        total_count, continuous_count, tire = wechat_mapping[name]
                        input_ws.cell(row=name_new_row + heart_rate_idx, column=19).value = total_count
                        input_ws.cell(row=name_new_row + heart_rate_idx, column=20).value = continuous_count
                        input_ws.cell(row=name_new_row + heart_rate_idx, column=21).value = tire
            else:
                print("[warning] %s not in file %s" % (name, input_file))
            a_name_row += 1
    # use student id to match in
    cleandata_wb = load_workbook(cleandata_file)
    cleandata_ws = cleandata_wb.active
    assert cleandata_ws.cell(row=1, column=1).value == "ID"
    assert cleandata_ws.cell(row=1, column=2).value == "pre_Age"
    assert cleandata_ws.cell(row=1, column=4).value == "pre_Yearofcollege"
    student_id_row = 2
    while cleandata_ws.cell(row=student_id_row, column=1).value is not None:
        student_id = str(cleandata_ws.cell(row=student_id_row, column=1).value)
        if student_id in id_to_row_mapping:
            name_row = id_to_row_mapping[student_id]
            name_new_row = get_name_new_row(name_row)
            input_ws.cell(row=name_new_row, column=3).value = cleandata_ws.cell(row=student_id_row, column=2).value
            input_ws.cell(row=name_new_row, column=4).value = cleandata_ws.cell(row=student_id_row, column=4).value
        student_id_row += 1
    input_wb.save(filename = output_file)

if __name__ == "__main__":
    main()
