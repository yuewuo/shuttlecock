from openpyxl import Workbook, load_workbook
import os, time

classes = ["class1and2"]

def main():
    for class_name in classes:
        append_data(class_name + ".xlsx", class_name + "_psycho_added.xlsx", "cleandata_all.xlsx")
        # exit(0)

def append_data(input_file, output_file, cleandata_file):
    input_wb = load_workbook(input_file)
    for sheet_name in input_wb.sheetnames:
        input_ws = input_wb[sheet_name]
        assert input_ws.cell(row=1, column=2).value == "姓名"
        assert input_ws.cell(row=1, column=1).value == "学号"
        # build name mapping
        name_row = 2
        id_to_row_mapping = { }
        while input_ws.cell(row=name_row, column=2).value is not None:
            name = input_ws.cell(row=name_row, column=2).value
            student_id = str(input_ws.cell(row=name_row, column=1).value)
            id_to_row_mapping[student_id] = name_row
            name_row += 1
        print(id_to_row_mapping)
        # find first none column
        first_none_column = 3
        while input_ws.cell(row=1, column=first_none_column).value is not None:
            first_none_column += 1
        # use student id to match in
        cleandata_wb = load_workbook(cleandata_file)
        cleandata_ws = cleandata_wb.active
        assert cleandata_ws.cell(row=1, column=1).value == "ID"
        copy_length = 0
        while cleandata_ws.cell(row=1, column=2+copy_length).value is not None:
            copy_length += 1
        # copy the head
        for copy_idx in range(copy_length):
            input_ws.cell(row=1, column=first_none_column+copy_idx).value = cleandata_ws.cell(row=1, column=2+copy_idx).value
        student_id_row = 2
        while cleandata_ws.cell(row=student_id_row, column=1).value is not None:
            student_id = str(cleandata_ws.cell(row=student_id_row, column=1).value)
            if student_id in id_to_row_mapping:
                name_row = id_to_row_mapping[student_id]
                for copy_idx in range(copy_length):
                    input_ws.cell(row=name_row, column=first_none_column+copy_idx).value = cleandata_ws.cell(row=student_id_row, column=2+copy_idx).value
            student_id_row += 1
        first_none_column += copy_length
    input_wb.save(filename = output_file)

if __name__ == '__main__':
    main()
