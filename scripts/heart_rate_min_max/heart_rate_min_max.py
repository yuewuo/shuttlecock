import xlrd, os, time, math
from openpyxl import Workbook, load_workbook
from string import digits

classes = ["class3", "class4", "class5"]

def main():
    for class_name in classes:
        heart_rate_min_max("../append_data/" + class_name + ".xlsx", "../analyze_heart_rate/data_" + class_name + ".xlsx", "min_max_" + class_name + ".xlsx")
        # exit(0)

def heart_rate_min_max(input_file, heart_rate_file, output_file):
    input_wb = load_workbook(input_file)
    input_ws = input_wb.active
    assert input_ws.cell(row=1, column=2).value == "姓名"
    # build name mapping
    name_row = 2
    name_to_row_mapping = { }
    while input_ws.cell(row=name_row, column=2).value is not None:
        name = input_ws.cell(row=name_row, column=2).value
        student_id = str(input_ws.cell(row=name_row, column=1).value)
        name_to_row_mapping[name] = name_row
        name_row += 1
    # find first none column
    first_none_column = 3
    while input_ws.cell(row=1, column=first_none_column).value is not None:
        first_none_column += 1
    # load data from heart rate file
    heart_rate_wb = load_workbook(heart_rate_file)
    for sheet_name in heart_rate_wb.sheetnames:
        heart_rate_ws = heart_rate_wb[sheet_name]
        assert heart_rate_ws.cell(row=1, column=2).value == "秒数"
        a_name_row = 2
        input_ws.cell(row=1, column=first_none_column).value = sheet_name + " 最小值"
        input_ws.cell(row=1, column=first_none_column + 1).value = sheet_name + " 最大值"
        input_ws.cell(row=1, column=first_none_column + 2).value = sheet_name + " 标准差"
        while heart_rate_ws.cell(row=a_name_row, column=1).value is not None:
            name = heart_rate_ws.cell(row=a_name_row, column=1).value.translate(str.maketrans('', '', digits))  # delete numbers in name
            if name in name_to_row_mapping:
                name_row = name_to_row_mapping[name]
                person_data = []
                for i in range(180):
                    value = heart_rate_ws.cell(row=a_name_row, column=3+i).value
                    if value is not None:
                        person_data.append(int(value))
                # compute min max dev
                person_min = min(person_data)
                person_max = max(person_data)
                average = float(sum(person_data)) / len(person_data)
                stddev = math.sqrt(sum([(e - average) ** 2 for e in person_data]) / len(person_data))
                # print(name, person_min, person_max, stddev)
                input_ws.cell(row=name_row, column=first_none_column).value = person_min
                input_ws.cell(row=name_row, column=first_none_column + 1).value = person_max
                input_ws.cell(row=name_row, column=first_none_column + 2).value = stddev
            else:
                print("[warning] %s not in file %s" % (name, input_file))
            a_name_row += 1
        first_none_column += 3
    input_wb.save(filename = output_file)

if __name__ == "__main__":
    main()
