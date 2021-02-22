import xlrd, os, time
from datetime import datetime, timedelta
from matplotlib import pyplot as plt
from openpyxl import Workbook

classes = {
    "data_class3": {
        "10.22.xls": (70, 420),
        "10.29.xls": (20, 540),
        "11.12.xls": (50, 330),
        "11.19.xls": (90, 400),
        "11.26.xls": (90, 420),
        "11.05.xls": (20, 400),
        "12.03.xls": (60, 310),
        "12.10.xls": (20, 280),
        "12.17.xls": (20, 290),
        "12.24.xls": (20, 320),
    },
    "data_class4": {
        "10.23.xls": (70, 400),
        "10.30.xls": (50, 390),
        "11.06.xls": (20, 450),
        "11.13.xls": (20, 360),
        "11.20.xls": (30, 380),
        "11.27.xls": (30, 290),
        "12.04.xls": (20, 280),
        "12.11.xls": (30, 290),
        "12.18.xls": (20, 320),
        "12.25.xls": (30, 260),
    },
    "data_class5": {
        "10.23.xls": (40, 340),
        "10.30.xls": (20, 280),
        "11.06.xls": (30, 310),
        "11.13.xls": (20, 280),
        "11.20.xls": (20, 245),
        "11.27.xls": (20, 300),
        "12.04.xls": (20, 290),
        "12.11.xls": (10, 280),
        "12.25.xls": (20, 280),
    },
    "data_graduate": {
        "10.21.xls": (20, 400),
        "10.28.xls": (140, 480),
        "11.04.xls": (200, 470),
        "11.11.xls": (70, 380),
        "11.18.xls": (50, 310),
        "11.25.xls": (20, 310),
        "12.02.xls": (60, 350),
        "12.09.xls": (30, 340),
        "12.16.xls": (80, 380),
        "12.23.xls": (120, 380),
    },
}

def main():
    # ## Step 1: manually generate `classes` variable for each class
    # # class_name = "data_class3"
    # # class_name = "data_class4"
    # # class_name = "data_class5"
    # class_name = "data_graduate"
    # file_list = os.listdir(class_name)
    # for filename in file_list:
    #     print("analyzing " + filename)
    #     filepath = os.path.abspath(os.path.join(class_name, filename))
    #     # plot single file to manually set `classes`, do this at first
    #     plot_single_file(filepath)
    
    ## Step 2: automatically split the data based on which is higher
    for class_name in classes:
        generate_splitted_file(class_name)

def generate_splitted_file(class_name):
    wb = Workbook()
    file_list = os.listdir(class_name)
    ws = None
    for filename in file_list:
        if ws is None:
            ws = wb.active
            ws.title = filename[:-4]
        else:
            ws = wb.create_sheet(title=filename[:-4])  # for example: 10.22
        filepath = os.path.abspath(os.path.join(class_name, filename))
        names, times, aligned_data = parse_single_file(read_single_file(filepath))
        former_start, latter_start = classes[class_name][filename]
        for i in range(180):
            ws.cell(row=1, column=i+3).value = i
        ws.cell(row=1, column=2).value = "秒数"
        for (idx, name) in enumerate(names):
            former_diff = get_diff_3min(aligned_data[idx], former_start)
            latter_diff = get_diff_3min(aligned_data[idx], latter_start)
            start_idx = former_start if former_diff > latter_diff else latter_start
            # print(name, former_diff, latter_diff, use_former)
            ws.cell(row=idx+2, column=1).value = name
            ws.cell(row=idx+2, column=2).value = "前半段" if former_diff > latter_diff else "后半段"
            for i in range(180):
                if start_idx + i < len(aligned_data[idx]) and aligned_data[idx][start_idx + i] > 0:
                    ws.cell(row=idx+2, column=i+3).value = aligned_data[idx][start_idx + i]
    wb.save(filename = os.path.abspath(class_name + ".xlsx"))

def get_diff_3min(data, start):
    begin = data[start]  # must exist
    if start + 180 < len(data):
        end = data[start + 180]
    else:
        end = data[-1]
    return end - begin

def plot_single_file(filepath):
    names, times, aligned_data = parse_single_file(read_single_file(filepath))
    times_x = [i for i in range(len(times))]
    plt.title(filepath)
    plt.xlabel("data point")
    plt.ylabel("heart rate")
    average_line = []
    for j in range(len(aligned_data[0])):
        cnt = 0
        value = 0
        for i in range(len(aligned_data)):
            if aligned_data[i][j] > 0:
                cnt += 1
                value += aligned_data[i][j]
        if cnt == 0:
            average_line.append(-1)
        else:
            average_line.append(value / cnt)
    plt.plot(times_x, average_line)
    # for i in range(len(aligned_data)):
    #     plt.plot(times, aligned_data[i])
    plt.show()

# return [name]: (time_values, heart_rate_values)
def read_single_file(filepath):
    workbook = xlrd.open_workbook(filepath)
    data = {}
    sorted_names = workbook.sheet_names()
    sorted_names.sort()
    for name in sorted_names:
        # name is something like '周亦凡26' or '陈仔佳23'
        table = workbook.sheet_by_name(name)
        assert table.ncols == 4, "format should be exactly the same, with data on 3rd and 4th column"
        time_values = table.col_values(2)
        heart_rate_values = table.col_values(3)
        assert time_values[0] == "时间", "please confirm the format of excel, 时间"
        assert heart_rate_values[0] == "原始心率", "please confirm the format of excel, 原始心率"
        assert len(time_values) == len(heart_rate_values), "row should be the same"
        time_values = [datetime.strptime(time, "%H:%M:%S")   for time in time_values[1:]]
        heart_rate_values = [int(heart_rate) for heart_rate in heart_rate_values[1:]]
        data[name] = (time_values, heart_rate_values)
    return data

def parse_single_file(data, verbose=False):
    # first find maximum start time and minimum end time
    minimum_start_time = None  # (name, time)
    maximum_end_time = None  # (name, time)
    for name in data:
        if minimum_start_time is None or data[name][0][0] < minimum_start_time[1]:
            minimum_start_time = (name, data[name][0][0])
        if maximum_end_time is None or data[name][0][-1] > maximum_end_time[1]:
            maximum_end_time = (name, data[name][0][-1])
    if verbose:
        print("minimum_start_time:", minimum_start_time[1].strftime("%H:%M:%S"), "from", minimum_start_time[0])
        print("maximum_end_time:", maximum_end_time[1].strftime("%H:%M:%S"), "from", maximum_end_time[0])
    names = []  # [name]
    time = minimum_start_time[1]
    times = [time]  # [time]
    while time < maximum_end_time[1]:
        time = time + timedelta(seconds=1)
        times.append(time)
    aligned_data = []  # [name_idx][time_idx] = heart_rate
    for name in data:
        names.append(name)
        row = []
        # sanity check
        for time in times:
            if time not in data[name][0]:
                # print("[warning]: " + name + " miss time point of " + time.strftime("%H:%M:%S") + ", use -1 for this time")
                row.append(-1)
            else:
                row.append(data[name][1][data[name][0].index(time)])
        aligned_data.append(row)
    return (names, times, aligned_data)

if __name__ == "__main__":
    main()
