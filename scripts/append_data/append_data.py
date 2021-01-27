from openpyxl import Workbook, load_workbook
import os, time
from string import digits

classes = ["class3", "class4", "class5"]

def main():
    for class_name in classes:
        append_data(class_name + ".xlsx", "../analyze_heart_rate/data_" + class_name + ".xlsx", class_name + "_appended.xlsx")
        # exit(0)

def append_data(input_file, to_be_appended, output_file, cleandata_file="cleandata_all.xlsx"):
    input_wb = load_workbook(input_file)
    appended_wb = load_workbook(to_be_appended)
    input_ws = input_wb.active
    assert input_ws.cell(row=1, column=2).value == "姓名"
    # build name mapping
    name_row = 2
    name_to_row_mapping = { }
    id_to_row_mapping = { }
    while input_ws.cell(row=name_row, column=2).value is not None:
        name = input_ws.cell(row=name_row, column=2).value
        student_id = str(input_ws.cell(row=name_row, column=1).value)
        name_to_row_mapping[name] = name_row
        id_to_row_mapping[student_id] = name_row
        name_row += 1
    # find first none column
    first_none_column = 3
    while input_ws.cell(row=1, column=first_none_column).value is not None:
        first_none_column += 1
    if cleandata_file is not None:
        # use student id to match in
        cleandata_wb = load_workbook(cleandata_file)
        cleandata_ws = cleandata_wb.active
        assert cleandata_ws.cell(row=1, column=1).value == "ID"
        copy_length = 0
        while cleandata_ws.cell(row=1, column=2+copy_length).value is not None:
            copy_length += 1
        # copy the head
        for copy_idx in range(copy_length):
            input_ws.cell(row=1, column=first_none_column+copy_idx).value = cleandata_ws.cell(row=1, column=2+copy_idx).value
        student_id_row = 2
        while cleandata_ws.cell(row=student_id_row, column=1).value is not None:
            student_id = str(cleandata_ws.cell(row=student_id_row, column=1).value)
            if student_id in id_to_row_mapping:
                name_row = id_to_row_mapping[student_id]
                for copy_idx in range(copy_length):
                    input_ws.cell(row=name_row, column=first_none_column+copy_idx).value = cleandata_ws.cell(row=student_id_row, column=2+copy_idx).value
            student_id_row += 1
        first_none_column += copy_length
    for sheet_name in appended_wb.sheetnames:
        input_ws.cell(row=1, column=first_none_column).value = "心率" + sheet_name
        for i in range(180):
            input_ws.cell(row=1, column=first_none_column+1+i).value = i
        appended_ws = appended_wb[sheet_name]
        a_name_row = 2
        while appended_ws.cell(row=a_name_row, column=1).value is not None:
            name = appended_ws.cell(row=a_name_row, column=1).value.translate(str.maketrans('', '', digits))  # delete numbers in name
            if name in name_to_row_mapping:
                name_row = name_to_row_mapping[name]
                input_ws.cell(row=name_row, column=first_none_column).value = appended_ws.cell(row=a_name_row, column=2).value
                for i in range(180):
                    input_ws.cell(row=name_row, column=first_none_column+1+i).value = appended_ws.cell(row=a_name_row, column=3+i).value
            else:
                print("[warning] %s not in file %s" % (name, input_file))
            a_name_row += 1
        first_none_column += 181
    input_wb.save(filename = output_file)

if __name__ == '__main__':
    main()
