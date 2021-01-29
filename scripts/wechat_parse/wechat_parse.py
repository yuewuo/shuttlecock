import re
from openpyxl import Workbook, load_workbook

def main():
    for class_name in ["class3", "class4", "class5"]:
        parsed_data = parse_txt_chat(class_name + ".txt")
        add_parsed_data_to_excel(class_name + ".xlsx", parsed_data, class_name + "_wechat.xlsx")
        # exit(0)

def add_parsed_data_to_excel(input_file, parsed_data, output_file):
    input_wb = load_workbook(input_file)
    input_ws = input_wb.active
    assert input_ws.cell(row=1, column=2).value == "姓名"
    # build name mapping
    name_row = 2
    name_to_row_mapping = { }
    while input_ws.cell(row=name_row, column=2).value is not None:
        name = input_ws.cell(row=name_row, column=2).value
        name_to_row_mapping[name] = name_row
        name_row += 1
    # find first none column
    first_none_column = 3
    while input_ws.cell(row=1, column=first_none_column).value is not None:
        first_none_column += 1
    for (tag, mapping) in parsed_data:
        input_ws.cell(row=1, column=first_none_column).value = tag
        input_ws.cell(row=1, column=first_none_column+1).value = "三次总数"
        input_ws.cell(row=1, column=first_none_column+2).value = "一次连续最多"
        input_ws.cell(row=1, column=first_none_column+3).value = "感觉"
        for name in mapping:
            if name not in name_to_row_mapping:
                print("[warning] name '%s' not in file %s" % (name, input_file))
            else:
                total_count, continuous_count, tire = mapping[name]
                name_row = name_to_row_mapping[name]
                input_ws.cell(row=name_row, column=first_none_column).value = str(total_count)
                input_ws.cell(row=name_row, column=first_none_column+1).value = str(continuous_count)
                input_ws.cell(row=name_row, column=first_none_column+2).value = tire
        first_none_column += 4
    input_wb.save(filename = output_file)

def parse_txt_chat(filename):
    with open(filename, "r", encoding="utf8") as f:
        lines = [line.strip() for line in f.readlines()]
    current_tag = None  # tag is something behind #!
    current_map = {}
    maps = []  #[(tag, map)]
    normal_pattern = re.compile(r'\d+\. ([^\d ]+)( *)(\d+)( +)(\d+)( *)([^\d ]*)')
    missing_single_pattern = re.compile(r'\d+\. ([^\d ]+)( *)(\d+)( *)([^\d ]*)')
    tire_pattern = re.compile(r'[累|稍累|累极了|正常]?')
    for line in lines:
        if line.startswith("#!"):
            if current_tag is not None:
                maps.append((current_tag, current_map))
            current_tag = line[2:].lstrip()
            current_map = {}
        elif line == "" or line.startswith("#"):
            pass  # do nothing, this is comment
        else:
            assert current_tag is not None, "reading line '%s' but missing current tag, try to add tag with #!" % line
            if re.match(r'\d+\. 言宜慢', line):
                pass
            elif normal_pattern.match(line):  # 覃斯欣 151 8 累
                matched = normal_pattern.match(line)
                name = matched.group(1)
                total_count = matched.group(3)
                continuous_count = matched.group(5)
                tire = matched.group(7)
                if tire_pattern.match(tire) is None:
                    print("[warning] tire pattern '%s' unknown in %s:" % (tire, filename), line)
                current_map[name] = (total_count, continuous_count, tire)
            elif missing_single_pattern.match(line):  # 黄筠茹 109 稍累
                matched = missing_single_pattern.match(line)
                name = matched.group(1)
                total_count = matched.group(3)
                tire = matched.group(5)
                if tire_pattern.match(tire) is None:
                    print("[warning] tire pattern '%s' unknown in %s:" % (tire, filename), line)
                current_map[name] = (total_count, "", tire)
            else:
                print("[warning] unhandled case in %s:" % filename, line)
    # print(maps)
    return maps

if __name__ == "__main__":
    main()
