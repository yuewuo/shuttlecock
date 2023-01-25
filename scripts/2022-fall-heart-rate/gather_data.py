from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Side, Border
from openpyxl.utils import get_column_letter
import os, time
from string import digits
from common import *
from pathlib import Path
import json
import matplotlib.pyplot as plt
import numpy as np
from tice.data import queries, Tice



def main():
    for (class_idx, (_, class_name)) in enumerate(class_names):
        gather_data(class_idx, class_name)
        # exit()

def gather_data(class_idx, class_name):
    # 体测数据（作为基准）
    tice = Tice(os.path.join(os.path.dirname(__file__), "tice", f"tice_{class_name}.xlsx"))
    # 心率数据
    class_info = class_infos[class_idx]
    hear_rate_data = { }
    for month, day, filename in class_info:
        date_str = f"{month}月{day}日"
        json_filepath = Path(class_name) / (filename[:-3] + "json")
        print(json_filepath)
        with open(json_filepath, "r", encoding="utf8") as f:
            data = json.load(f)
            for student in data['students']:
                name = student["name"]
                if name not in hear_rate_data:
                    hear_rate_data[name] = { }
                assert date_str not in hear_rate_data[name]
                assert len(student['truncations']) <= 2
                selected_truncations = []
                for truncation in student['truncations']:
                    select = True
                    if truncation[173] < 85.0:
                        select = False
                    if truncation[178] < 112:
                        select = False
                    if truncation[94] < 102:
                        select = False
                    if truncation[281] > 177:
                        select = False
                    if truncation[336] is None or truncation[336] > 163:
                        select = False
                    if select:
                        selected_truncations.append(truncation)
                hear_rate_data[name][date_str] = selected_truncations
    # style
    thin = Side(border_style="thin", color="000000")
    double = Side(border_style="double", color="ff0000")
    border = Border(top=thin, left=double, right=thin, bottom=thin)
    # 生成合并文件头
    gathered_wb = Workbook()
    gathered_ws = gathered_wb.active
    title_count = 0
    row_count = 1 + len(tice.student_names)
    tice_column_bias = 1
    for title_idx, title in enumerate(tice.titles):
        gathered_ws.cell(row=1, column=tice_column_bias+title_idx).value = title
        title_count += 1
    fill = PatternFill("solid", fgColor="ffd699")
    for row in range(1, row_count + 1):
        gathered_ws.cell(row=row, column=tice_column_bias).fill = fill
        gathered_ws.cell(row=row, column=tice_column_bias).border = border
    query_column_bias = []
    for query_idx in range(2):
        query_column_bias.append(1 + title_count)
        query = queries[query_idx]
        for title_idx, title in enumerate(query.titles):
            gathered_ws.cell(row=1, column=query_column_bias[query_idx]+title_idx).value = title
            title_count += 1
        fill = PatternFill("solid", fgColor="ff9999" if query_idx == 0 else "99ddff")
        for row in range(1, row_count + 1):
            gathered_ws.cell(row=row, column=query_column_bias[query_idx]).fill = fill
            gathered_ws.cell(row=row, column=query_column_bias[query_idx]).border = border
    heart_rate_column_bias = 1 + title_count
    for heart_rate_idx, (month, day, filename) in enumerate(class_info):
        date_str = f"{month}月{day}日"
        column_bias = heart_rate_column_bias + (2 * record_length + 1) * heart_rate_idx
        gathered_ws.cell(row=1, column=column_bias).value = date_str
        column_width = gathered_ws.column_dimensions[get_column_letter(column_bias)].width * 0.3
        for i in range(record_length):
            gathered_ws.cell(row=1, column=column_bias+1+i).value = f"{i}"
            gathered_ws.cell(row=1, column=column_bias+1+record_length+i).value = f"{i}"
            gathered_ws.column_dimensions[get_column_letter(column_bias+1+i)].width = column_width
            gathered_ws.column_dimensions[get_column_letter(column_bias+1+record_length+i)].width = column_width
        fill = PatternFill("solid", fgColor="ccff33")
        fill_1 = PatternFill("solid", fgColor="d699ff")
        fill_2 = PatternFill("solid", fgColor="d4d4aa")
        for row in range(1, row_count + 1):
            gathered_ws.cell(row=row, column=column_bias).fill = fill
            gathered_ws.cell(row=row, column=column_bias).border = border
            gathered_ws.cell(row=row, column=column_bias + 1).fill = fill_1
            gathered_ws.cell(row=row, column=column_bias + 1).border = border
            gathered_ws.cell(row=row, column=column_bias + 1 + record_length).fill = fill_2
            gathered_ws.cell(row=row, column=column_bias + 1 + record_length).border = border
    # 填充数据
    for stu_idx, name in enumerate(tice.student_names):
        row = 2 + stu_idx
        line = tice.data[stu_idx]
        for i, value in enumerate(line):
            gathered_ws.cell(row=row, column=tice_column_bias+i).value = value
        for query_idx in range(2):
            query = queries[query_idx]
            line = query.get_by_name(name)
            if line is not None:
                for i, value in enumerate(line):
                    gathered_ws.cell(row=row, column=query_column_bias[query_idx]+i).value = value
        if name in hear_rate_data:
            for heart_rate_idx, (month, day, filename) in enumerate(class_info):
                date_str = f"{month}月{day}日"
                if date_str in hear_rate_data[name]:
                    truncations = hear_rate_data[name][date_str]
                    for trunc_idx, truncation in enumerate(truncations):
                        column_bias = heart_rate_column_bias + (2 * record_length + 1) * heart_rate_idx + trunc_idx * record_length + 1
                        for i, value in enumerate(truncation):
                            gathered_ws.cell(row=row, column=column_bias+i).value = value
    # 保存
    gathered_wb.save(filename = f"gathered_{class_name}.xlsx")


if __name__ == "__main__":
    main()
