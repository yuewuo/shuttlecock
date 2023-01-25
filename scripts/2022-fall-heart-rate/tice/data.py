from openpyxl import Workbook, load_workbook
import os, time
import matplotlib.pyplot as plt
import numpy as np


directory = os.path.dirname(__file__)

# get 心理量表
query_filenames = ["中华毽课程心理量表学期前问卷_136.xlsx", "中华毽课程心理量表学期末问卷_147.xlsx"]
queries = []

class Query:
    def __init__(self, query_filepath):
        input_wb = load_workbook(query_filepath)
        input_ws = input_wb[input_wb.sheetnames[0]]
        name_column = 7
        assert input_ws.cell(row=1, column=name_column).value == "1、姓名"
        # build name mapping
        name_row = 2
        name_to_row_mapping = { }
        while input_ws.cell(row=name_row, column=name_column).value is not None:
            name = input_ws.cell(row=name_row, column=name_column).value
            if name in name_to_row_mapping:
                print(f"[warning] duplicate name ({name}) in query: {query_filepath}")
            name_to_row_mapping[name] = name_row
            name_row += 1
        # find first none column
        titles = []
        column_indices = []
        for i in range(1, 100):
            title = input_ws.cell(row=1, column=i).value
            if title is not None:
                titles.append(title)
                column_indices.append(i)
        # gather data in a map from student name to row
        data = { }
        for name in name_to_row_mapping:
            row = name_to_row_mapping[name]
            line = []
            for column in column_indices:
                line.append(input_ws.cell(row=row, column=column).value)
            data[name] = line
        self.data = data
        self.titles = titles
    def get_by_name(self, name):
        if name not in self.data:
            return None
        return self.data[name]

for query_filename in query_filenames:
    query_filepath = os.path.join(directory, query_filename)
    query = Query(query_filepath)
    queries.append(query)

class Tice:
    def __init__(self, filepath):
        input_wb = load_workbook(filepath)
        input_ws = input_wb[input_wb.sheetnames[0]]
        name_column = 3
        assert input_ws.cell(row=3, column=5).value == "体测成绩（20分）"
        assert input_ws.cell(row=4, column=5).value == "得分"
        assert input_ws.cell(row=4, column=6).value == "肺活量"
        assert input_ws.cell(row=3, column=name_column).value == "姓名"
        # build name mapping
        name_row = 5
        name_to_row_mapping = { }
        id_to_row_mapping = { }
        student_names = []
        while input_ws.cell(row=name_row, column=name_column).value is not None:
            name = input_ws.cell(row=name_row, column=name_column).value
            student_id = str(input_ws.cell(row=name_row, column=2).value)
            name_to_row_mapping[name] = name_row
            id_to_row_mapping[student_id] = name_row
            student_names.append(name)
            name_row += 1
        # find first none column
        titles = []
        column_indices = []
        for i in range(1, 30):
            title = input_ws.cell(row=3, column=i).value
            if title == "体测成绩（20分）" or (title is None and input_ws.cell(row=4, column=i).value is not None):
                title = input_ws.cell(row=4, column=i).value
            if title is not None:
                titles.append(title)
                column_indices.append(i)
        # gather data in a map from student name to row
        data = []
        for row in range(5, name_row):
            line = []
            for column in column_indices:
                line.append(input_ws.cell(row=row, column=column).value)
            data.append(line)
        self.name_to_row_mapping = name_to_row_mapping
        self.id_to_row_mapping = id_to_row_mapping
        self.data = data
        self.titles = titles
        self.student_names = student_names
    def get_by_name(self, name):
        if name not in self.name_to_row_mapping:
            return None
        return self.data[self.name_to_row_mapping[name]]
    def get_by_student_id(self, student_id):
        if student_id not in self.id_to_row_mapping:
            return None
        return self.data[self.id_to_row_mapping[student_id]]
